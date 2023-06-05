import pandas as pd
import glob
import openpyxl
from fpdf import FPDF
from pathlib import Path

filepaths = glob.glob("invoices/*.xlsx")

for filepath in filepaths:
    # For reading a xlsx file which is an Excel file we need read_excel
    df= pd.read_excel(filepath, sheet_name="Sheet 1")

    pdf = FPDF(orientation="P", unit="mm", format="A4")
    pdf.add_page()

    filename = Path(filepath).stem
    invoice_nr, date = filename.split("-")

    pdf.set_font(family="Times", size=16, style="B")
    pdf.cell(w=50, h=8, txt=f"Invoice nr.{invoice_nr}", ln=1)

    pdf.set_font(family="Times", size=16, style="B")
    pdf.cell(w=50, h=8, txt=f"Date :{date}", ln=1)

    # Add a header
    a = list(df.columns)
    columns = [columns.replace("_"," ").title() for columns in a]
    pdf.set_font(family="Times", size=10, style="B")
    pdf.set_text_color(80, 80, 80)
    pdf.cell(w=30, h=8, txt=columns[0], border=1)
    pdf.cell(w=70, h=8, txt=columns[1], border=1)
    pdf.cell(w=30, h=8, txt=columns[2], border=1)
    pdf.cell(w=30, h=8, txt=columns[3], border=1)
    pdf.cell(w=30, h=8, txt=columns[4], border=1, ln=1)

    for index, row in df.iterrows():
        pdf.set_font(family="Times", size=10)
        pdf.set_text_color(80, 80, 80)
        pdf.cell(w=30, h=8, txt=str(row["product_id"]), border=1)
        pdf.cell(w=70, h=8, txt=str(row["product_name"]), border=1)
        pdf.cell(w=30, h=8, txt=str(row["amount_purchased"]), border=1)
        pdf.cell(w=30, h=8, txt=str(row["price_per_unit"]), border=1)
        pdf.cell(w=30, h=8, txt=str(row["total_price"]), border=1, ln=1)


    a = df["total_price"]
    b=a.sum()
    pdf.set_text_color(80, 80, 80)
    pdf.cell(w=30, h=8, txt="", border=1)
    pdf.cell(w=70, h=8, txt="", border=1)
    pdf.cell(w=30, h=8, txt="", border=1)
    pdf.cell(w=30, h=8, txt="", border=1)
    pdf.cell(w=30, h=8, txt=str(b), border=1, ln=1)

    # Add total sum sentence
    pdf.set_font(family="Times", size=10, style="B")
    pdf.cell(w=30, h=8, txt=f"The total price is {b}", ln=1)

    # Add company name and logo
    pdf.set_font(family="Times", size=14, style="B")
    pdf.cell(w=30, h=8, txt=f"PythonHow")
    pdf.image("pythonhow.png", w=10)


    pdf.output(f"PDFs/{filename}.pdf")







